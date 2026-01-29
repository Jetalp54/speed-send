from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import logging
import csv
import io
import re

from app.database import get_db
from app.models import DataList
from app.schemas import DataListCreate, DataListUpdate, DataListResponse

router = APIRouter(prefix="/data-lists", tags=["data-lists"])
logger = logging.getLogger(__name__)

# Email validation regex
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

def is_valid_email(email: str) -> bool:
    """Validate email format"""
    if not email or not isinstance(email, str):
        return False
    return bool(EMAIL_REGEX.match(email.strip()))

def detect_email_column(rows: List[List[str]]) -> tuple[int, str]:
    """
    Detect which column contains emails.
    Returns: (column_index, column_name)
    """
    if not rows or len(rows) < 2:
        return (0, "Unknown")
    
    headers = rows[0]
    
    # Strategy 1: Check headers for common email column names
    email_keywords = ['email', 'e-mail', 'mail', 'contact', 'recipient']
    for idx, header in enumerate(headers):
        if any(keyword in header.lower() for keyword in email_keywords):
            logger.info(f"Detected email column by header: '{header}' at index {idx}")
            return (idx, header)
    
    # Strategy 2: Scan first 10 rows for @ symbols
    max_scan = min(10, len(rows))
    column_email_counts = {}
    
    for col_idx in range(len(headers)):
        email_count = 0
        for row_idx in range(1, max_scan):  # Skip header row
            if col_idx < len(rows[row_idx]):
                cell_value = str(rows[row_idx][col_idx]).strip()
                if '@' in cell_value and is_valid_email(cell_value):
                    email_count += 1
        column_email_counts[col_idx] = email_count
    
    # Find column with most emails
    if column_email_counts:
        best_col = max(column_email_counts.items(), key=lambda x: x[1])
        col_idx, count = best_col
        if count > 0:
            col_name = headers[col_idx] if col_idx < len(headers) else f"Column {col_idx + 1}"
            logger.info(f"Detected email column by content: '{col_name}' at index {col_idx} ({count} emails found)")
            return (col_idx, col_name)
    
    # Default: assume first column
    logger.warning("Could not detect email column, defaulting to first column")
    return (0, headers[0] if headers else "Column 1")

def parse_csv_file(file_content: bytes) -> List[List[str]]:
    """Parse CSV file content into rows"""
    try:
        text = file_content.decode('utf-8')
        reader = csv.reader(io.StringIO(text))
        return list(reader)
    except UnicodeDecodeError:
        # Try with different encoding
        text = file_content.decode('latin-1')
        reader = csv.reader(io.StringIO(text))
        return list(reader)

def parse_excel_file(file_content: bytes) -> List[List[str]]:
    """Parse Excel file content into rows"""
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        
        rows = []
        for row in sheet.iter_rows(values_only=True):
            # Convert row tuple to list of strings, handling None values
            rows.append([str(cell) if cell is not None else '' for cell in row])
        
        return rows
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

@router.post("/upload-file")
async def upload_data_list_file(file: UploadFile = File(...)):
    """
    Upload and parse a data list file (CSV or Excel).
    Returns extracted emails and metadata.
    """
    try:
        logger.info(f"Uploading file: {file.filename}")
        
        # Read file content
        file_content = await file.read()
        
        # Detect file type and parse
        filename_lower = file.filename.lower()
        
        if filename_lower.endswith('.csv'):
            rows = parse_csv_file(file_content)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            rows = parse_excel_file(file_content)
        else:
            raise HTTPException(
                status_code=400, 
                detail="Unsupported file format. Please upload CSV (.csv) or Excel (.xlsx, .xls) files."
            )
        
        if not rows or len(rows) < 1:
            raise HTTPException(status_code=400, detail="File appears to be empty")
        
        # Detect email column
        email_col_idx, email_col_name = detect_email_column(rows)
        
        # Extract emails from detected column
        emails = []
        for row_idx, row in enumerate(rows[1:], start=2):  # Skip header, start at row 2 for logging
            if email_col_idx < len(row):
                email = str(row[email_col_idx]).strip()
                if email and is_valid_email(email):
                    emails.append(email)
                elif email:  # Non-empty but invalid
                    logger.debug(f"Skipping invalid email at row {row_idx}: {email}")
        
        # Remove duplicates while preserving order
        unique_emails = []
        seen = set()
        for email in emails:
            email_lower = email.lower()
            if email_lower not in seen:
                seen.add(email_lower)
                unique_emails.append(email)
        
        logger.info(f"Extracted {len(unique_emails)} unique valid emails from {len(rows)-1} rows")
        
        return {
            "success": True,
            "emails": unique_emails,
            "total": len(unique_emails),
            "detected_column": email_col_name,
            "total_rows": len(rows) - 1,  # Exclude header
            "preview": unique_emails[:10]  # First 10 for preview
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to upload file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

@router.get("", response_model=List[DataListResponse])
async def list_data_lists(db: Session = Depends(get_db)):
    """List all data lists"""
    try:
        logger.info("Fetching data lists...")
        data_lists = db.query(DataList).all()
        logger.info(f"Found {len(data_lists)} data lists")
        return data_lists
    except Exception as e:
        logger.error(f"Failed to fetch data lists: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("", response_model=DataListResponse)
async def create_data_list(data_list: DataListCreate, db: Session = Depends(get_db)):
    """Create a new data list"""
    try:
        logger.info(f"Creating data list: {data_list.name}")
        
        # Calculate total recipients
        total_recipients = len(data_list.recipients) if data_list.recipients else 0
        
        db_data_list = DataList(
            name=data_list.name,
            description=data_list.description,
            recipients=data_list.recipients,
            total_recipients=total_recipients,
            list_type=data_list.list_type
        )
        
        db.add(db_data_list)
        db.commit()
        db.refresh(db_data_list)
        
        logger.info(f"Successfully created data list: {db_data_list.name} (ID: {db_data_list.id})")
        return db_data_list
        
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to create data list: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{list_id}", response_model=DataListResponse)
async def get_data_list(list_id: int, db: Session = Depends(get_db)):
    """Get a specific data list"""
    data_list = db.query(DataList).filter(DataList.id == list_id).first()
    if not data_list:
        raise HTTPException(status_code=404, detail="Data list not found")
    return data_list

@router.put("/{list_id}", response_model=DataListResponse)
async def update_data_list(
    list_id: int,
    data_list_update: DataListUpdate,
    db: Session = Depends(get_db)
):
    """Update a data list"""
    data_list = db.query(DataList).filter(DataList.id == list_id).first()
    if not data_list:
        raise HTTPException(status_code=404, detail="Data list not found")
    
    try:
        # Update fields
        update_data = data_list_update.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(data_list, field, value)
        
        # Recalculate total recipients if recipients were updated
        if 'recipients' in update_data and update_data['recipients'] is not None:
            data_list.total_recipients = len(update_data['recipients'])
        
        db.commit()
        db.refresh(data_list)
        
        logger.info(f"Successfully updated data list: {data_list.name}")
        return data_list
        
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to update data list: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{list_id}")
async def delete_data_list(list_id: int, db: Session = Depends(get_db)):
    """Delete a data list"""
    data_list = db.query(DataList).filter(DataList.id == list_id).first()
    if not data_list:
        raise HTTPException(status_code=404, detail="Data list not found")
    
    try:
        db.delete(data_list)
        db.commit()
        
        logger.info(f"Successfully deleted data list: {data_list.name}")
        return {"message": "Data list deleted successfully"}
        
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to delete data list: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/search/{query}", response_model=List[DataListResponse])
async def search_data_lists(query: str, db: Session = Depends(get_db)):
    """Search data lists by name or description"""
    try:
        logger.info(f"Searching data lists for: {query}")
        
        data_lists = db.query(DataList).filter(
            DataList.name.ilike(f"%{query}%") | 
            DataList.description.ilike(f"%{query}%")
        ).all()
        
        logger.info(f"Found {len(data_lists)} data lists matching query")
        return data_lists
        
    except Exception as e:
        logger.error(f"Failed to search data lists: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{list_id}/add-recipients")
async def add_recipients_to_list(
    list_id: int,
    recipients: List[str],
    db: Session = Depends(get_db)
):
    """Add recipients to an existing data list"""
    try:
        logger.info(f"Adding {len(recipients)} recipients to data list {list_id}")
        
        data_list = db.query(DataList).filter(DataList.id == list_id).first()
        if not data_list:
            raise HTTPException(status_code=404, detail="Data list not found")
        
        # Add new recipients to existing list
        existing_recipients = data_list.recipients or []
        new_recipients = [email for email in recipients if email not in existing_recipients]
        updated_recipients = existing_recipients + new_recipients
        
        data_list.recipients = updated_recipients
        data_list.total_recipients = len(updated_recipients)
        
        db.commit()
        db.refresh(data_list)
        
        logger.info(f"Successfully added {len(new_recipients)} new recipients to data list")
        return {
            "message": f"Successfully added {len(new_recipients)} new recipients",
            "total_recipients": data_list.total_recipients,
            "new_recipients_added": len(new_recipients)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to add recipients to data list: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{list_id}/remove-recipients")
async def remove_recipients_from_list(
    list_id: int,
    recipients: List[str],
    db: Session = Depends(get_db)
):
    """Remove recipients from a data list"""
    try:
        logger.info(f"Removing {len(recipients)} recipients from data list {list_id}")
        
        data_list = db.query(DataList).filter(DataList.id == list_id).first()
        if not data_list:
            raise HTTPException(status_code=404, detail="Data list not found")
        
        # Remove recipients from existing list
        existing_recipients = data_list.recipients or []
        updated_recipients = [email for email in existing_recipients if email not in recipients]
        
        data_list.recipients = updated_recipients
        data_list.total_recipients = len(updated_recipients)
        
        db.commit()
        db.refresh(data_list)
        
        removed_count = len(existing_recipients) - len(updated_recipients)
        logger.info(f"Successfully removed {removed_count} recipients from data list")
        return {
            "message": f"Successfully removed {removed_count} recipients",
            "total_recipients": data_list.total_recipients,
            "recipients_removed": removed_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to remove recipients from data list: {e}")
        raise HTTPException(status_code=500, detail=str(e))